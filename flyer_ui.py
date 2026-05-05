"""
flyer_ui.py — Modulul de UI pentru generarea flyere.

Expune funcția render() care construiește toate tab-urile flyer.
Apelat din app.py când utilizatorul alege "Flyer" în sidebar.
"""
import streamlit as st
import pandas as pd
import io
from flyer_engine import (
    DEFAULT_STYLE, DEFAULT_ANTET, DEFAULT_OBSERVATII,
    img_bytes_to_b64, generate_html, build_pptx,
)


# ── INIT SESSION STATE ─────────────────────────────────────────
def init_state():
    if "antet" not in st.session_state:
        st.session_state.antet = dict(DEFAULT_ANTET)
    if "antet_image_bytes" not in st.session_state:
        st.session_state.antet_image_bytes = None
    if "antet_image_name" not in st.session_state:
        st.session_state.antet_image_name = None
    if "style" not in st.session_state:
        st.session_state.style = dict(DEFAULT_STYLE)
    if "observatii" not in st.session_state:
        st.session_state.observatii = list(DEFAULT_OBSERVATII)
    # Numele celor 3 coloane de atribute - utilizatorul le poate redenumi
    if "attr_names" not in st.session_state:
        st.session_state.attr_names = ["Adâncime", "Înălțime", "Finisaj"]
    if "produse_df" not in st.session_state:
        # Folosim numele atributelor curente
        a1, a2, a3 = st.session_state.attr_names
        st.session_state.produse_df = pd.DataFrame([
            {
                "Cod": "2012959", "Denumire": "Set sertar Vision Top",
                a1: "350mm", a2: "89mm", a3: "Alb",
                "Preț": 74, "Monedă": "LEI", "UM": "buc",
                "Reducere (%)": None, "Preț redus": None,
                "Badge": "ALB", "Imagine badge": "",
                "Imagine": "89-alb.jpg", "Text preț": "Preț cu TVA inclus",
                "Text bulină": "NOU",
            },
            {
                "Cod": "2012976", "Denumire": "Set sertar Vision Top",
                a1: "350mm", a2: "185mm", a3: "Alb",
                "Preț": 94, "Monedă": "LEI", "UM": "buc",
                "Reducere (%)": None, "Preț redus": None,
                "Badge": "ALB", "Imagine badge": "",
                "Imagine": "185-alb.jpg", "Text preț": "Preț cu TVA inclus",
                "Text bulină": "",
            },
            {
                "Cod": "2012978", "Denumire": "Set sertar Vision Top",
                a1: "400mm", a2: "185mm", a3: "Alb",
                "Preț": 99, "Monedă": "LEI", "UM": "buc",
                "Reducere (%)": 15, "Preț redus": 84,
                "Badge": "OFERTĂ", "Imagine badge": "",
                "Imagine": "185-alb.jpg", "Text preț": "Preț cu TVA inclus",
                "Text bulină": "",
            },
        ])
    if "images" not in st.session_state:
        st.session_state.images = {}  # {filename: bytes}




def render():
    """Construiește interfața pentru generare flyere."""
    init_state()


    # ── HEADER ─────────────────────────────────────────────────────
    # Hero banner corporate
    st.markdown("""
    <div class="brand-hero">
        <div class="brand-hero-cat">METALCOM · PRINT TOOLS</div>
        <h1 class="brand-hero-title">Generator Flyer</h1>
        <div class="brand-hero-subtitle">
            Configurează datele, stilul și layout-ul, apoi generează HTML și PPTX.
        </div>
    </div>
    """, unsafe_allow_html=True)


    # ── TABURI ─────────────────────────────────────────────────────
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "🏢 Antet",
        "📦 Produse",
        "🎨 Stil",
        "📝 Observații",
        "🚀 Generează",
    ])


    # ══════════════════════════════════════════════════════════════
    # TAB 1: ANTET
    # ══════════════════════════════════════════════════════════════
    with tab1:
        st.subheader("Date firmă și texte din header")

        col_a, col_b = st.columns(2)

        with col_a:
            st.session_state.antet["Nume companie"] = st.text_input(
                "Nume companie", st.session_state.antet["Nume companie"])
            st.session_state.antet["Categorie (sus)"] = st.text_input(
                "Categorie (text mic deasupra titlului)", st.session_state.antet["Categorie (sus)"])
            st.session_state.antet["Titlu produs (mare)"] = st.text_input(
                "Titlu produs (text mare)", st.session_state.antet["Titlu produs (mare)"])
            st.session_state.antet["Text promoțional mic"] = st.text_input(
                "Text promoțional mic (dreapta sus)", st.session_state.antet["Text promoțional mic"])
            st.session_state.antet["Text promoțional mare"] = st.text_input(
                "Text promoțional mare (dreapta sus)", st.session_state.antet["Text promoțional mare"])

        with col_b:
            st.session_state.antet["Telefon"] = st.text_input(
                "Telefon", st.session_state.antet["Telefon"])
            st.session_state.antet["Email"] = st.text_input(
                "Email", st.session_state.antet["Email"])
            st.session_state.antet["Website 1"] = st.text_input(
                "Website 1", st.session_state.antet["Website 1"])
            st.session_state.antet["Website 2"] = st.text_input(
                "Website 2", st.session_state.antet["Website 2"])

        st.session_state.antet["Footer legal"] = st.text_area(
            "Footer legal (jos pe pagină)",
            st.session_state.antet["Footer legal"],
            height=80,
        )

        st.divider()
        st.subheader("Logo antet")
        logo = st.file_uploader(
            "Încarcă imaginea de antet (logo + date firmă)",
            type=["jpg", "jpeg", "png"],
            help="Imaginea ocupă întreaga lățime a paginii. Recomandare: 1920x388px sau aspect similar.",
        )
        if logo:
            st.session_state.antet_image_bytes = logo.getvalue()
            st.session_state.antet_image_name = logo.name
            st.success(f"✓ Logo încărcat: {logo.name}")
            st.image(logo, width=600)
        elif st.session_state.antet_image_bytes:
            st.info(f"ℹ️ Folosesc logo-ul încărcat anterior: {st.session_state.antet_image_name}")
            st.image(st.session_state.antet_image_bytes, width=600)

        st.divider()
        with st.expander("⚠️ Golire date antet"):
            st.caption(
                "Șterge toate datele din antet (texte + logo). Datele din alte tab-uri "
                "(produse, stil, observații) NU sunt afectate."
            )
            confirm_clear_antet = st.checkbox(
                "Confirm că vreau să șterg datele din antet",
                key="confirm_clear_antet",
            )
            if st.button("🗑️ Golește antet", disabled=not confirm_clear_antet, key="btn_clear_antet"):
                st.session_state.antet = {k: "" for k in DEFAULT_ANTET.keys()}
                st.session_state.antet_image_bytes = None
                st.session_state.antet_image_name = None
                st.success("✓ Antet golit!")
                st.rerun()


    # ══════════════════════════════════════════════════════════════
    # TAB 2: PRODUSE
    # ══════════════════════════════════════════════════════════════
    with tab2:
        st.subheader("Lista produselor")

        # ── Editare nume coloane atribute ──
        with st.expander("⚙️ Configurare nume atribute (3 coloane personalizabile)", expanded=False):
            st.caption(
                "Cele 3 coloane de atribute pot fi redenumite. Numele apare ca etichetă pe flyer "
                "(ex: »Adâncime: 450mm«). Schimbă o coloană în »Culoare«, »Volum«, »Material« etc."
            )
            c1, c2, c3 = st.columns(3)
            with c1:
                new_attr1 = st.text_input("Atribut 1", st.session_state.attr_names[0], key="attr1_input")
            with c2:
                new_attr2 = st.text_input("Atribut 2", st.session_state.attr_names[1], key="attr2_input")
            with c3:
                new_attr3 = st.text_input("Atribut 3", st.session_state.attr_names[2], key="attr3_input")

            new_names = [new_attr1.strip() or "Atribut 1",
                         new_attr2.strip() or "Atribut 2",
                         new_attr3.strip() or "Atribut 3"]

            if new_names != st.session_state.attr_names:
                # Redenumim coloanele în DataFrame păstrând valorile
                old_names = st.session_state.attr_names
                rename_map = {old_names[i]: new_names[i] for i in range(3) if old_names[i] != new_names[i]}
                if rename_map:
                    st.session_state.produse_df = st.session_state.produse_df.rename(columns=rename_map)
                    st.session_state.attr_names = new_names
                    st.success(f"✓ Coloane redenumite: {', '.join(new_names)}")
                    st.rerun()

        # ── Import / Export Excel ──
        with st.expander("📥 Import/Export Excel", expanded=False):
            col_imp, col_exp = st.columns(2)

            with col_imp:
                st.markdown("**Import din Excel**")
                uploaded_xlsx = st.file_uploader(
                    "Încarcă fișier Excel cu produse",
                    type=["xlsx"],
                    key="xlsx_import",
                    help="Fișierul trebuie să aibă pe primul rând antetul cu numele coloanelor.",
                )
                if uploaded_xlsx:
                    try:
                        df_imported = pd.read_excel(uploaded_xlsx, sheet_name=0)
                        # Detectăm numele coloanelor de atribute (3, 4, 5 - dacă există)
                        cols = list(df_imported.columns)
                        if len(cols) >= 5:
                            # Detectăm dacă coloana 3, 4, 5 sunt atribute (după numele lor sau poziție)
                            new_attrs = [str(c).strip() for c in cols[2:5]]
                            # Dacă au nume, le folosim
                            if all(new_attrs):
                                st.session_state.attr_names = new_attrs
                        st.session_state.produse_df = df_imported
                        st.success(f"✓ Importat: {len(df_imported)} produse")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Eroare la import: {e}")

            with col_exp:
                st.markdown("**Descarcă model Excel**")
                st.caption("Fișier model gata-completat care poate fi folosit ca template de import.")

                # Generăm model Excel în memorie
                from io import BytesIO
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment

                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Produse"

                    a1, a2, a3 = st.session_state.attr_names
                    headers = [
                        "Cod", "Denumire", a1, a2, a3,
                        "Preț", "Monedă", "UM",
                        "Reducere (%)", "Preț redus",
                        "Badge", "Imagine badge", "Imagine", "Text preț", "Text bulină"
                    ]
                    for i, h in enumerate(headers, start=1):
                        c = ws.cell(row=1, column=i, value=h)
                        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                        c.fill = PatternFill("solid", fgColor="A11319")
                        c.alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[1].height = 28

                    # Adăugăm câteva rânduri de exemplu
                    example_rows = [
                        ["2012959", "Set sertar Vision Top", "350mm", "89mm", "Alb",
                         74, "LEI", "buc", None, None,
                         "ALB", "", "89-alb.jpg", "Preț cu TVA inclus", "NOU"],
                        ["2012976", "Set sertar Vision Top", "350mm", "185mm", "Alb",
                         94, "LEI", "buc", None, None,
                         "ALB", "", "185-alb.jpg", "Preț cu TVA inclus", ""],
                        ["2012978", "Set sertar Vision Top", "400mm", "185mm", "Alb",
                         99, "LEI", "buc", 15, 84,
                         "OFERTĂ", "", "185-alb.jpg", "Preț cu TVA inclus", ""],
                        ["2012962", "Set sertar Vision Top", "400mm", "89mm", "Alb",
                         78, "LEI", "buc", None, None,
                         "IP44", "waterdrop.png", "89-alb.jpg", "Preț cu TVA inclus", ""],
                    ]
                    for r, row in enumerate(example_rows, start=2):
                        for col_i, val in enumerate(row, start=1):
                            ws.cell(row=r, column=col_i, value=val)

                    # Lățimi coloane
                    widths = [12, 22, 14, 14, 14, 10, 10, 10, 11, 12, 13, 18, 22, 24, 14]
                    for i, w in enumerate(widths, start=1):
                        ws.column_dimensions[chr(64+i) if i <= 26 else 'A' + chr(64+i-26)].width = w

                    buf = BytesIO()
                    wb.save(buf)
                    buf.seek(0)

                    st.download_button(
                        "📥 Descarcă model_produse.xlsx",
                        data=buf.getvalue(),
                        file_name="model_produse.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.warning(f"Nu pot genera model Excel: {e}")

        st.caption("Editează direct în tabel. Adaugă rânduri cu butonul + de la sfârșitul tabelului.")

        # Configurare dinamică pentru cele 3 coloane atribute
        a1, a2, a3 = st.session_state.attr_names

        edited_df = st.data_editor(
            st.session_state.produse_df,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "Cod":          st.column_config.TextColumn("Cod", width="small"),
                "Denumire":     st.column_config.TextColumn("Denumire", width="medium"),
                a1:             st.column_config.TextColumn(a1, width="small"),
                a2:             st.column_config.TextColumn(a2, width="small"),
                a3:             st.column_config.TextColumn(a3, width="small"),
                "Preț":         st.column_config.NumberColumn("Preț", format="%.0f"),
                "Monedă":       st.column_config.TextColumn("Mon.", width="small"),
                "UM":           st.column_config.TextColumn("UM", width="small"),
                "Reducere (%)": st.column_config.NumberColumn("Red%", format="%d"),
                "Preț redus":   st.column_config.NumberColumn("P.red.", format="%.0f"),
                "Badge":        st.column_config.TextColumn("Badge", width="small"),
                "Imagine badge":st.column_config.TextColumn("Img badge", width="small"),
                "Imagine":      st.column_config.TextColumn("Imagine", width="medium"),
                "Text preț":    st.column_config.TextColumn("Text preț", width="medium"),
                "Text bulină":  st.column_config.TextColumn("Bulină", width="small"),
            },
            key="produse_editor",
        )
        st.session_state.produse_df = edited_df

        st.info(
            f"💡 **Coloane importante:**\n"
            f"- **{a1}, {a2}, {a3}** — atribute personalizabile (modifică numele în secțiunea ⚙️ de sus)\n"
            f"- **Imagine** și **Imagine badge** — numele fișierelor pe care le încarci mai jos\n"
            f"- **Text bulină** — text custom (ex: NOU, HIT, -30%). Dacă e gol, se folosește Reducere %\n"
            f"- **Badge** — text afișat pe casetă (ex: ALB, IP44)"
        )

        st.divider()
        st.subheader("Imagini produse și pictograme badge")

        uploaded_images = st.file_uploader(
            "Încarcă imaginile produselor (numele fișierelor trebuie să corespundă coloanelor »Imagine« și »Imagine badge«)",
            type=["jpg", "jpeg", "png", "webp"],
            accept_multiple_files=True,
            key="prod_images",
        )

        if uploaded_images:
            for img in uploaded_images:
                st.session_state.images[img.name] = img.getvalue()

        if st.session_state.images:
            st.success(f"✓ {len(st.session_state.images)} imagini disponibile")

            with st.expander("Vezi imaginile încărcate"):
                cols = st.columns(4)
                for idx, (name, data) in enumerate(sorted(st.session_state.images.items())):
                    with cols[idx % 4]:
                        st.image(data, caption=name, use_column_width=True)

            if st.button("🗑️ Șterge toate imaginile"):
                st.session_state.images = {}
                st.rerun()

        st.divider()
        with st.expander("⚠️ Golire date produse"):
            st.caption(
                "Șterge toate produsele din tabel (rămâne tabelul gol cu antetul de coloane) "
                "și/sau toate imaginile încărcate."
            )
            col_g1, col_g2 = st.columns(2)
            with col_g1:
                confirm_clear_prod = st.checkbox(
                    "Confirm ștergere produse",
                    key="confirm_clear_prod",
                )
                if st.button("🗑️ Golește tabel produse",
                             disabled=not confirm_clear_prod, key="btn_clear_prod"):
                    a1, a2, a3 = st.session_state.attr_names
                    empty_cols = [
                        "Cod", "Denumire", a1, a2, a3,
                        "Preț", "Monedă", "UM",
                        "Reducere (%)", "Preț redus",
                        "Badge", "Imagine badge", "Imagine", "Text preț", "Text bulină",
                    ]
                    st.session_state.produse_df = pd.DataFrame(columns=empty_cols)
                    st.success("✓ Tabel produse golit!")
                    st.rerun()
            with col_g2:
                confirm_clear_imgs = st.checkbox(
                    "Confirm ștergere imagini",
                    key="confirm_clear_imgs",
                )
                if st.button("🗑️ Șterge toate imaginile produselor",
                             disabled=not confirm_clear_imgs, key="btn_clear_imgs"):
                    st.session_state.images = {}
                    st.success("✓ Imagini șterse!")
                    st.rerun()


    # ══════════════════════════════════════════════════════════════
    # TAB 3: STIL
    # ══════════════════════════════════════════════════════════════
    with tab3:
        st.subheader("Personalizare vizuală")
        st.caption("Modifică culorile, fonturile și dimensiunile elementelor de pe flyer.")

        SAFE_FONTS = [
            "Montserrat", "Roboto", "Open Sans", "Lato", "Poppins",
            "Arial", "Helvetica", "Verdana", "Tahoma",
            "Times New Roman", "Georgia", "Trebuchet MS", "Courier New",
        ]

        # ── Brand colors ──
        st.markdown("### 🎨 Culori brand")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.session_state.style["Culoare principală (roșu)"] = st.color_picker(
                "Culoare principală",
                "#" + st.session_state.style["Culoare principală (roșu)"],
            ).lstrip("#").upper()
        with col2:
            st.session_state.style["Culoare secundară (gri)"] = st.color_picker(
                "Culoare secundară",
                "#" + st.session_state.style["Culoare secundară (gri)"],
            ).lstrip("#").upper()
        with col3:
            st.session_state.style["Culoare fundal pagină"] = st.color_picker(
                "Fundal pagină",
                "#" + st.session_state.style["Culoare fundal pagină"],
            ).lstrip("#").upper()
        with col4:
            st.session_state.style["Culoare footer (jos pagină)"] = st.color_picker(
                "Culoare footer",
                "#" + st.session_state.style["Culoare footer (jos pagină)"],
            ).lstrip("#").upper()

        st.divider()

        # ── Hero (titlul mare) ──
        st.markdown("### 📰 Titlul mare (HERO)")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.session_state.style["Font titlu produs (mare)"] = st.selectbox(
                "Font titlu",
                SAFE_FONTS,
                index=SAFE_FONTS.index(st.session_state.style["Font titlu produs (mare)"])
                      if st.session_state.style["Font titlu produs (mare)"] in SAFE_FONTS else 0,
            )
        with col2:
            st.session_state.style["Dimensiune titlu produs (mare)"] = st.slider(
                "Dimensiune titlu",
                16, 48, int(st.session_state.style["Dimensiune titlu produs (mare)"]),
            )
        with col3:
            st.session_state.style["Culoare titlu produs (mare)"] = st.color_picker(
                "Culoare titlu",
                "#" + st.session_state.style["Culoare titlu produs (mare)"],
            ).lstrip("#").upper()

        st.divider()

        # ── Denumire produs ──
        st.markdown("### 📦 Denumire produs (în casetă)")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.session_state.style["Font denumire produs"] = st.selectbox(
                "Font",
                SAFE_FONTS,
                index=SAFE_FONTS.index(st.session_state.style["Font denumire produs"])
                      if st.session_state.style["Font denumire produs"] in SAFE_FONTS else 0,
                key="font_denumire",
            )
        with col2:
            st.session_state.style["Dimensiune denumire produs"] = st.slider(
                "Dimensiune", 12, 36, int(st.session_state.style["Dimensiune denumire produs"]),
                key="dim_denumire",
            )
        with col3:
            st.session_state.style["Culoare denumire produs"] = st.color_picker(
                "Culoare",
                "#" + st.session_state.style["Culoare denumire produs"],
                key="color_denumire",
            ).lstrip("#").upper()

        st.divider()

        # ── Preț ──
        st.markdown("### 💰 Preț")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.session_state.style["Font preț"] = st.selectbox(
                "Font preț",
                SAFE_FONTS,
                index=SAFE_FONTS.index(st.session_state.style["Font preț"])
                      if st.session_state.style["Font preț"] in SAFE_FONTS else 0,
                key="font_pret",
            )
        with col2:
            st.session_state.style["Dimensiune preț"] = st.slider(
                "Dim. preț", 20, 50, int(st.session_state.style["Dimensiune preț"]),
                key="dim_pret",
            )
        with col3:
            st.session_state.style["Dimensiune »LEI« preț"] = st.slider(
                "Dim. LEI/€", 10, 28, int(st.session_state.style["Dimensiune »LEI« preț"]),
                key="dim_lei",
            )
        with col4:
            st.session_state.style["Dimensiune UM (/ buc)"] = st.slider(
                "Dim. UM", 8, 24, int(st.session_state.style["Dimensiune UM (/ buc)"]),
                key="dim_um",
            )

        st.divider()

        # ── Bulină ──
        st.markdown("### 🔴 Bulină reducere")
        col1, col2 = st.columns(2)
        with col1:
            st.session_state.style["Diametru bulină reducere (px)"] = st.slider(
                "Diametru bulină", 36, 100, int(st.session_state.style["Diametru bulină reducere (px)"]),
            )
        with col2:
            st.session_state.style["Dimensiune cifră reducere"] = st.slider(
                "Dimensiune text bulină", 14, 40, int(st.session_state.style["Dimensiune cifră reducere"]),
            )

        st.divider()

        # ── Badge ──
        st.markdown("### 🏷️ Badge produs")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.session_state.style["Culoare fundal badge"] = st.color_picker(
                "Fundal badge",
                "#" + st.session_state.style["Culoare fundal badge"],
            ).lstrip("#").upper()
        with col2:
            st.session_state.style["Culoare text badge"] = st.color_picker(
                "Text badge",
                "#" + st.session_state.style["Culoare text badge"],
            ).lstrip("#").upper()
        with col3:
            st.session_state.style["Dimensiune font badge (px)"] = st.slider(
                "Mărime font", 8, 18, int(st.session_state.style["Dimensiune font badge (px)"]),
            )

        st.divider()

        # ── Reset ──
        if st.button("🔄 Resetează stilul la valorile implicite"):
            st.session_state.style = dict(DEFAULT_STYLE)
            st.success("Stil resetat!")
            st.rerun()


    # ══════════════════════════════════════════════════════════════
    # TAB 4: OBSERVAȚII
    # ══════════════════════════════════════════════════════════════
    with tab4:
        st.subheader("Observații afișate în footer-ul flyerului")

        obs_text = "\n".join(st.session_state.observatii)
        new_obs = st.text_area(
            "Una pe linie (fără ▪ — se adaugă automat)",
            obs_text,
            height=200,
        )
        st.session_state.observatii = [
            line.strip() for line in new_obs.split("\n") if line.strip()
        ]

        st.caption(f"✓ {len(st.session_state.observatii)} observații")


    # ══════════════════════════════════════════════════════════════
    # TAB 5: GENERARE
    # ══════════════════════════════════════════════════════════════
    with tab5:
        st.subheader("Selectează layout-ul și generează")

        col_left, col_right = st.columns([2, 1])

        with col_left:
            layout_options = {
                "1": "📄 1 produs/pagină (fișa produs)",
                "2": "📑 2 produse/pagină (orizontale lungi)",
                "3": "📋 3 produse/pagină (medii)",
                "3a": "🌟 3a — 1 mare sus + 2 jos",
                "4": "▣ 4 produse/pagină (2×2 pătrate)",
                "5": "▥ 5 produse/pagină (4 + 1 lat)",
                "6": "⊞ 6 produse/pagină (default)",
            }
            layout_key = st.radio(
                "Layout",
                options=list(layout_options.keys()),
                format_func=lambda x: layout_options[x],
                index=6,  # default 6
                horizontal=True,
            )

            # Convertim "3a" rămâne string, restul devin int
            per_page = layout_key if layout_key == "3a" else int(layout_key)

            chunk_size = 3 if layout_key == "3a" else int(layout_key)
            total_products = len(st.session_state.produse_df)
            total_pages = (total_products + chunk_size - 1) // chunk_size

            st.metric(
                "Estimare",
                f"{total_pages} pagini",
                f"{total_products} produse · {chunk_size}/pagină",
            )

        with col_right:
            st.write("**Opțiuni**")
            embed_images_in_html = st.checkbox(
                "Include imagini în HTML",
                value=True,
                help="Bifat: HTML conține imaginile (fișier mai mare dar portabil). Debifat: imaginile sunt linkuri externe."
            )

        st.divider()

        # ── Acțiuni ──
        col1, col2, col3 = st.columns([1, 1, 1])

        with col1:
            generate_btn = st.button("🚀 Generează HTML + PPTX", type="primary", use_container_width=True)

        if generate_btn:
            # Convertim tabelul în format așteptat de motor
            a1, a2, a3 = st.session_state.attr_names
            attr_cols = [a1, a2, a3]

            produse = []
            for _, row in st.session_state.produse_df.iterrows():
                cod = str(row.get("Cod", "")).strip()
                if not cod or cod == "nan":
                    continue

                attrs = []
                for col_attr in attr_cols:
                    val = row.get(col_attr)
                    if val is not None and str(val).strip() and str(val).strip() != "nan":
                        attrs.append({"label": col_attr, "value": str(val).strip()})

                def safe_str(v):
                    s = str(v).strip() if v is not None else ""
                    return "" if s == "nan" else s

                def safe_num(v):
                    if v is None or str(v).strip() in ("", "nan"):
                        return None
                    try:
                        return float(v)
                    except (ValueError, TypeError):
                        return None

                produse.append({
                    "cod": cod,
                    "denumire": safe_str(row.get("Denumire", "")),
                    "attrs": attrs,
                    "pret": safe_num(row.get("Preț")),
                    "moneda": safe_str(row.get("Monedă", "")),
                    "um": safe_str(row.get("UM", "")),
                    "reducere": safe_num(row.get("Reducere (%)")),
                    "pret_redus": safe_num(row.get("Preț redus")),
                    "badge_text": safe_str(row.get("Badge", "")),
                    "badge_image": safe_str(row.get("Imagine badge", "")),
                    "imagine": safe_str(row.get("Imagine", "")),
                    "tva_text": safe_str(row.get("Text preț", "")),
                    "bulina_text": safe_str(row.get("Text bulină", "")),
                })

            if not produse:
                st.error("⚠️ Nu există produse în tabel!")
            else:
                with st.spinner("Generez flyerul..."):
                    # Pregătim cache-ul de imagini base64 pentru HTML
                    img_cache = {}
                    for name, data in st.session_state.images.items():
                        img_cache[name] = img_bytes_to_b64(data)

                    antet_b64 = ""
                    if st.session_state.antet_image_bytes:
                        antet_b64 = img_bytes_to_b64(
                            st.session_state.antet_image_bytes, max_width=1920
                        )

                    # Generăm HTML
                    html_content = generate_html(
                        produse=produse,
                        antet=st.session_state.antet,
                        observatii=st.session_state.observatii,
                        style=st.session_state.style,
                        img_cache=img_cache,
                        antet_b64=antet_b64,
                        per_page=per_page,
                    )

                    # Generăm PPTX
                    pptx_bytes = build_pptx(
                        produse=produse,
                        antet=st.session_state.antet,
                        observatii=st.session_state.observatii,
                        style=st.session_state.style,
                        image_files=st.session_state.images,
                        antet_image_bytes=st.session_state.antet_image_bytes,
                        per_page=per_page,
                    )

                    # Salvăm în session pentru download
                    st.session_state.last_html = html_content
                    st.session_state.last_pptx = pptx_bytes
                    st.session_state.last_filename = "Flyer_generat"

                st.success(f"✅ Flyer generat: {total_pages} pagini cu {len(produse)} produse")

        # ── Butoane download ──
        if "last_html" in st.session_state and st.session_state.last_html:
            st.divider()
            st.subheader("📥 Descarcă fișierele generate")

            st.info(
                "💡 **Cum obții PDF-ul:** descarcă HTML-ul, deschide-l în browser, "
                "apasă butonul roșu »Descarcă PDF« din colțul dreapta-jos (sau Ctrl+P), "
                "apoi alege »Save as PDF« și bifează »Background graphics«."
            )

            col_dl1, col_dl2 = st.columns(2)

            with col_dl1:
                st.download_button(
                    "📄 Descarcă HTML",
                    data=st.session_state.last_html.encode("utf-8"),
                    file_name=f"{st.session_state.last_filename}.html",
                    mime="text/html",
                    use_container_width=True,
                    help="HTML cu buton încorporat pentru salvare ca PDF.",
                )

            with col_dl2:
                if st.session_state.get("last_pptx"):
                    st.download_button(
                        "📊 Descarcă PPTX (Canva)",
                        data=st.session_state.last_pptx,
                        file_name=f"{st.session_state.last_filename}.pptx",
                        mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                        use_container_width=True,
                        help="Importabil în Canva, PowerPoint sau Google Slides.",
                    )
                else:
                    st.warning("⚠️ PPTX indisponibil")
                    st.caption("Lipsește python-pptx.")

            # ── Preview HTML ──
            with st.expander("👁️ Preview flyer (HTML)", expanded=True):
                st.components.v1.html(
                    st.session_state.last_html,
                    height=900,
                    scrolling=True,
                )
