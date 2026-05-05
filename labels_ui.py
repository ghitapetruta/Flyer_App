"""
labels_ui.py — Modulul de UI pentru generarea etichetelor.

Expune funcția render() care construiește toate tab-urile pentru etichete.
Apelat din app.py când utilizatorul alege "Etichete" în sidebar.
"""
import streamlit as st
import pandas as pd
from io import BytesIO

from labels_engine import (
    DEFAULT_LABEL_CONFIG, generate_html, logo_to_b64,
)


# ══════════════════════════════════════════════════════════════
# INIT SESSION STATE
# ══════════════════════════════════════════════════════════════
def init_state():
    if "labels_logo_bytes" not in st.session_state:
        st.session_state.labels_logo_bytes = None
    if "labels_logo_name" not in st.session_state:
        st.session_state.labels_logo_name = None
    if "labels_config" not in st.session_state:
        st.session_state.labels_config = dict(DEFAULT_LABEL_CONFIG)
    if "labels_df" not in st.session_state:
        st.session_state.labels_df = pd.DataFrame([
            {"Cod": "2006147", "Denumire": "MANER UA 124-160-L1", "UM": "Buc",
             "Preț": 27, "Reducere (%)": 20, "Preț redus": 21.6},
            {"Cod": "2006148", "Denumire": "MANER UA 124-192-L1", "UM": "Buc",
             "Preț": 30, "Reducere (%)": 20, "Preț redus": 23.99},
            {"Cod": "2005953", "Denumire": "MANER RE 23-0096-G7", "UM": "Buc",
             "Preț": 21.4, "Reducere (%)": None, "Preț redus": None},
        ])


# ══════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════
def _build_model_xlsx() -> bytes:
    """Generează un fișier Excel model pentru import."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return b""

    wb = Workbook()
    ws = wb.active
    ws.title = "Etichete"

    headers = ["Cod", "Denumire", "UM", "Preț", "Reducere (%)", "Preț redus"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="C8102E")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    example_rows = [
        ["2006147", "MANER UA 124-160-L1", "Buc", 27, 20, 21.6],
        ["2006148", "MANER UA 124-192-L1", "Buc", 30, 20, 23.99],
        ["2005953", "MANER RE 23-0096-G7", "Buc", 21.4, None, None],
        ["2006177", "MANER UN 59/160 04", "Buc", 19.4, None, None],
    ]
    for r, row in enumerate(example_rows, start=2):
        for col_i, val in enumerate(row, start=1):
            ws.cell(row=r, column=col_i, value=val)

    widths = [12, 28, 8, 10, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _df_to_products(df: pd.DataFrame) -> list[dict]:
    """Convertește DataFrame-ul în lista de produse pentru engine."""
    products = []
    for _, row in df.iterrows():
        cod = str(row.get("Cod", "")).strip()
        if not cod or cod == "nan":
            continue

        def safe_str(v):
            s = str(v).strip() if v is not None else ""
            return "" if s == "nan" else s

        def safe_num(v):
            if v is None or str(v).strip() in ("", "nan"):
                return None
            try:
                return float(v)
            except (ValueError, TypeError):
                return None

        products.append({
            "cod": cod,
            "denumire": safe_str(row.get("Denumire", "")),
            "um": safe_str(row.get("UM", "")) or "buc",
            "pret": safe_num(row.get("Preț")),
            "reducere": safe_num(row.get("Reducere (%)")),
            "pret_redus": safe_num(row.get("Preț redus")),
        })
    return products


# ══════════════════════════════════════════════════════════════
# RENDER
# ══════════════════════════════════════════════════════════════
def render():
    """Construiește interfața pentru generarea etichetelor."""
    init_state()

    # Hero banner corporate
    st.markdown("""
    <div class="brand-hero">
        <div class="brand-hero-cat">METALCOM · PRINT TOOLS</div>
        <h1 class="brand-hero-title">Generator Etichete</h1>
        <div class="brand-hero-subtitle">
            Etichete A4 cu logo, cod, denumire și preț — layout flexibil 1-5 coloane.
        </div>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2, tab3, tab4 = st.tabs([
        "⚙️ Setări",
        "📦 Produse",
        "📐 Layout",
        "🚀 Generează",
    ])

    # ──────────────────────────────────────────────────────────
    # TAB 1: SETĂRI
    # ──────────────────────────────────────────────────────────
    with tab1:
        st.subheader("Logo (opțional)")
        st.caption(
            "Logo-ul se afișează în colțul dreapta-sus al fiecărei etichete. "
            "Recomandare: imagine cu fundal alb sau transparent (PNG)."
        )
        logo = st.file_uploader(
            "Încarcă logo",
            type=["jpg", "jpeg", "png"],
            key="labels_logo_upload",
        )
        if logo:
            st.session_state.labels_logo_bytes = logo.getvalue()
            st.session_state.labels_logo_name = logo.name
            st.success(f"✓ Logo încărcat: {logo.name}")
            st.image(logo, width=200)
        elif st.session_state.labels_logo_bytes:
            st.info(f"ℹ️ Logo curent: {st.session_state.labels_logo_name}")
            st.image(st.session_state.labels_logo_bytes, width=200)

            if st.button("🗑️ Șterge logo", key="del_label_logo"):
                st.session_state.labels_logo_bytes = None
                st.session_state.labels_logo_name = None
                st.rerun()

        st.divider()
        st.subheader("Texte afișate pe etichete")
        col_t1, col_t2 = st.columns(2)
        with col_t1:
            st.session_state.labels_config["moneda"] = st.text_input(
                "Monedă (afișată după preț)",
                st.session_state.labels_config.get("moneda", "Lei"),
            )
        with col_t2:
            st.session_state.labels_config["tva_text"] = st.text_input(
                "Text TVA (lângă monedă)",
                st.session_state.labels_config.get("tva_text", "TVA Inclus"),
            )

        st.session_state.labels_config["color_red"] = st.color_picker(
            "Culoare reducere (badge + preț redus)",
            st.session_state.labels_config.get("color_red", "#C8102E"),
        )

    # ──────────────────────────────────────────────────────────
    # TAB 2: PRODUSE
    # ──────────────────────────────────────────────────────────
    with tab2:
        st.subheader("Lista produselor pentru etichete")

        with st.expander("📥 Import/Export Excel", expanded=False):
            col_imp, col_exp = st.columns(2)

            with col_imp:
                st.markdown("**Import din Excel**")
                uploaded_xlsx = st.file_uploader(
                    "Încarcă fișier Excel cu produse",
                    type=["xlsx"],
                    key="labels_xlsx_import",
                    help=(
                        "Fișierul trebuie să aibă pe primul rând antetul cu coloanele: "
                        "Cod, Denumire, UM, Preț, Reducere, Preț redus."
                    ),
                )
                if uploaded_xlsx:
                    try:
                        df_imported = pd.read_excel(uploaded_xlsx, sheet_name=0)
                        # Normalizează numele coloanelor — acceptă și "Pret" fără diacritică
                        rename_map = {}
                        for col in df_imported.columns:
                            cl = str(col).strip().lower()
                            if cl == "pret":
                                rename_map[col] = "Preț"
                            elif cl == "pret redus":
                                rename_map[col] = "Preț redus"
                            elif cl == "reducere":
                                rename_map[col] = "Reducere (%)"
                            elif cl == "reducere (%)":
                                rename_map[col] = "Reducere (%)"
                        if rename_map:
                            df_imported = df_imported.rename(columns=rename_map)
                        st.session_state.labels_df = df_imported
                        st.success(f"✓ Importat: {len(df_imported)} produse")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Eroare la import: {e}")

            with col_exp:
                st.markdown("**Descarcă model Excel**")
                st.caption("Fișier model gata-completat ce poate fi folosit ca template.")
                model_bytes = _build_model_xlsx()
                if model_bytes:
                    st.download_button(
                        "📥 Descarcă model_etichete.xlsx",
                        data=model_bytes,
                        file_name="model_etichete.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                else:
                    st.warning("Nu pot genera modelul (lipsește openpyxl).")

        st.caption("Editează direct în tabel. Adaugă rânduri cu butonul + de la sfârșitul tabelului.")

        edited_df = st.data_editor(
            st.session_state.labels_df,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "Cod":          st.column_config.TextColumn("Cod", width="small"),
                "Denumire":     st.column_config.TextColumn("Denumire", width="large"),
                "UM":           st.column_config.TextColumn("UM", width="small"),
                "Preț":         st.column_config.NumberColumn("Preț", format="%.2f"),
                "Reducere (%)": st.column_config.NumberColumn("Red%", format="%d"),
                "Preț redus":   st.column_config.NumberColumn("P.redus", format="%.2f"),
            },
            key="labels_editor",
        )
        st.session_state.labels_df = edited_df

        st.info(
            "💡 **Reducerea și prețul redus sunt opționale.** "
            "Dacă lași gol »Preț redus«, eticheta afișează doar prețul normal. "
            "Dacă completezi, eticheta va arăta prețul vechi (tăiat) + procent + preț nou."
        )

        st.divider()
        with st.expander("⚠️ Golire date"):
            st.caption("Șterge toate produsele din tabel.")
            confirm = st.checkbox(
                "Confirm că vreau să șterg produsele",
                key="labels_confirm_clear",
            )
            if st.button("🗑️ Golește tabel", disabled=not confirm, key="labels_btn_clear"):
                st.session_state.labels_df = pd.DataFrame(
                    columns=["Cod", "Denumire", "UM", "Preț", "Reducere (%)", "Preț redus"]
                )
                st.success("✓ Tabel golit!")
                st.rerun()

    # ──────────────────────────────────────────────────────────
    # TAB 3: LAYOUT
    # ──────────────────────────────────────────────────────────
    with tab3:
        st.subheader("Configurare layout pagină A4")
        st.caption("Definește numărul de coloane și rânduri, precum și dimensiunile etichetelor.")

        col_l1, col_l2 = st.columns(2)
        with col_l1:
            st.session_state.labels_config["cols"] = st.slider(
                "Coloane",
                min_value=1, max_value=5,
                value=int(st.session_state.labels_config.get("cols", 2)),
                help="Numărul de etichete pe orizontală. Mai multe coloane = etichete mai înguste.",
            )
        with col_l2:
            st.session_state.labels_config["rows"] = st.slider(
                "Rânduri",
                min_value=1, max_value=12,
                value=int(st.session_state.labels_config.get("rows", 8)),
                help="Numărul de etichete pe verticală.",
            )

        cols_n = st.session_state.labels_config["cols"]
        rows_n = st.session_state.labels_config["rows"]
        per_page = cols_n * rows_n
        st.metric("Etichete pe pagină", f"{per_page}", f"{cols_n} col × {rows_n} rânduri")

        st.divider()
        st.subheader("Dimensiuni (mm)")
        col_d1, col_d2, col_d3 = st.columns(3)
        with col_d1:
            st.session_state.labels_config["label_height_mm"] = st.number_input(
                "Înălțime etichetă (mm)",
                min_value=15.0, max_value=100.0, step=1.0,
                value=float(st.session_state.labels_config.get("label_height_mm", 31)),
            )
        with col_d2:
            st.session_state.labels_config["page_margin_mm"] = st.number_input(
                "Margine pagină (mm)",
                min_value=0.0, max_value=30.0, step=0.5,
                value=float(st.session_state.labels_config.get("page_margin_mm", 8)),
            )
        with col_d3:
            st.session_state.labels_config["gap_h_mm"] = st.number_input(
                "Spațiu orizontal (mm)",
                min_value=0.0, max_value=10.0, step=0.5,
                value=float(st.session_state.labels_config.get("gap_h_mm", 4)),
            )

        st.session_state.labels_config["gap_v_mm"] = st.number_input(
            "Spațiu vertical (mm)",
            min_value=0.0, max_value=10.0, step=0.5,
            value=float(st.session_state.labels_config.get("gap_v_mm", 2)),
        )

        # Validare: încape totul pe pagină?
        total_h_needed = (
            rows_n * st.session_state.labels_config["label_height_mm"]
            + (rows_n - 1) * st.session_state.labels_config["gap_v_mm"]
        )
        page_h_avail = 297 - 2 * st.session_state.labels_config["page_margin_mm"]
        if total_h_needed > page_h_avail:
            st.warning(
                f"⚠️ Înălțimea totală necesară ({total_h_needed:.1f}mm) depășește "
                f"spațiul disponibil pe pagină ({page_h_avail:.1f}mm). "
                f"Reduceți înălțimea etichetei sau numărul de rânduri."
            )

    # ──────────────────────────────────────────────────────────
    # TAB 4: GENEREAZĂ
    # ──────────────────────────────────────────────────────────
    with tab4:
        st.subheader("Generare etichete")

        cols_n = int(st.session_state.labels_config["cols"])
        rows_n = int(st.session_state.labels_config["rows"])
        per_page = cols_n * rows_n
        total_products = len(st.session_state.labels_df)
        # Numără doar produse cu Cod completat
        valid_count = sum(
            1 for _, row in st.session_state.labels_df.iterrows()
            if str(row.get("Cod", "")).strip() not in ("", "nan")
        )
        total_pages = (valid_count + per_page - 1) // per_page if per_page else 0

        col_m1, col_m2, col_m3 = st.columns(3)
        with col_m1:
            st.metric("Produse valide", valid_count)
        with col_m2:
            st.metric("Pe pagină", per_page)
        with col_m3:
            st.metric("Pagini estimate", total_pages)

        st.divider()

        if st.button("🚀 Generează HTML cu etichete",
                     type="primary", use_container_width=True):
            products = _df_to_products(st.session_state.labels_df)
            if not products:
                st.error("⚠️ Nu există produse valide în tabel!")
            else:
                with st.spinner("Generez etichetele..."):
                    logo_b64 = logo_to_b64(st.session_state.labels_logo_bytes)
                    html = generate_html(
                        products=products,
                        cfg=st.session_state.labels_config,
                        logo_b64=logo_b64,
                    )
                    st.session_state.labels_last_html = html

                st.success(f"✅ Generat: {valid_count} etichete pe {total_pages} pagini")

        if st.session_state.get("labels_last_html"):
            st.divider()
            st.subheader("📥 Descarcă fișierul")

            st.info(
                "💡 **Cum obții PDF-ul:** descarcă HTML-ul, deschide-l în browser, "
                "apasă »⬇ Descarcă PDF« din colțul dreapta-sus (sau Ctrl+P), "
                "apoi alege »Save as PDF« și bifează »Background graphics«."
            )

            st.download_button(
                "📄 Descarcă HTML cu etichete",
                data=st.session_state.labels_last_html.encode("utf-8"),
                file_name="etichete.html",
                mime="text/html",
                use_container_width=True,
            )

            with st.expander("👁️ Preview etichete (HTML)", expanded=True):
                st.components.v1.html(
                    st.session_state.labels_last_html,
                    height=900,
                    scrolling=True,
                )
